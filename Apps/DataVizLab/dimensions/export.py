from tempfile import NamedTemporaryFile

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def export_to_xlsx(request, titles, dataFrames, file_name):
    wb = Workbook()
    wb.remove(wb.active)  # Elimina la hoja de trabajo predeterminada

    # Definir los estilos de relleno
    fill_inpar = PatternFill(start_color="A3CAFF", end_color="A3CAFF", fill_type="solid")
    fill_par = PatternFill(start_color="D6EEF7", end_color="D6EEF7", fill_type="solid")
    fill_header = PatternFill(start_color="62A0F5", end_color="62A0F5", fill_type="solid")

    for i in range(len(titles)):
        title = titles[i]
        df = dataFrames[i]

        ws = wb.create_sheet(title)

        if title == 'Segmentación por género':
            rows = dataframe_to_rows(df, index=True, header=False)
        else:
            rows = dataframe_to_rows(df, index=False)

        # Escribir los datos y aplicar el color a los encabezados
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Aplicar estilo a los encabezados
                    cell.fill = fill_header

        # Aplicar colores alternados a las filas (excepto los encabezados)
        for r_idx, row in enumerate(
                ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
            fill = fill_par if r_idx % 2 == 0 else fill_inpar
            for cell in row:
                cell.fill = fill

    # Usar NamedTemporaryFile para crear un archivo temporal
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        response = HttpResponse(tmp.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
        return response
